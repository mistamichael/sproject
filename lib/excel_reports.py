"""
Excel Report Generator for Gantt Charts and Resource Lists
===========================================================

Generates Excel worksheets with Gantt charts and resource allocation diagrams.
"""
import csv
import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Tuple, Any
import configparser

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import models - try relative import first, then absolute
try:
    from .models import Project
    from .models.reports import GanttReport, ResourceListReport
    from .models.cpm import CPMResult
    from .utils import format_time_value_auto, add_workdays, interpolate_color
except (ImportError, ValueError):
    from models import Project
    from models.reports import GanttReport, ResourceListReport
    from models.cpm import CPMResult
    from utils import format_time_value_auto, add_workdays, interpolate_color


def load_excel_export_config(cfg_dir: Path) -> Dict[str, Any]:
    """
    Lädt Excel-Export-Konfiguration aus excel_export.cfg und defaults.cfg.

    Args:
        cfg_dir: Verzeichnis mit Konfigurationsdateien

    Returns:
        Dictionary mit Konfigurationswerten (alle Sektionen)
    """
    config = configparser.ConfigParser()

    # Lade zuerst excel_export.cfg, dann defaults.cfg als Fallback
    excel_config_file = cfg_dir / "excel_export.cfg"
    defaults_config_file = cfg_dir / "defaults.cfg"

    # Default-Werte für alle Sektionen
    defaults = {
        'Gantt_resolution': {
            'threshold_hours_min': 60,
            'threshold_halfday_min': 240,
            'work_start_hour': 7,
            'work_end_hour': 20,
            'working_days': 'mon,tue,wed,thu,fri',
        },
        'GanttChart': {
            'color_start': '8B7AB8',
            'color_end': '4472C4',
            'critical_border_color': 'FF0000',
            'slack_bar_transparency': 180,
            'bar_height': 20,
            'timeline_header_height': 25,
            'col_width_id': 12,
            'col_width_name': 20,
            'col_width_start': 12,
            'col_width_end': 12,
            'col_width_effort': 10,
        },
        'ResourceList': {
            'col_width_resource': 20,
            'col_width_task': 25,
            'col_width_start': 12,
            'col_width_end': 12,
            'col_width_duration': 10,
            'color_person': '8B7AB8',
            'color_machine': 'FFA500',
            'color_material': '90EE90',
            'weekend_background_color': 'F0F0F0',
        }
    }

    # Lade excel_export.cfg falls vorhanden
    if excel_config_file.exists():
        config.read(excel_config_file, encoding='utf-8')
    # Fallback auf defaults.cfg
    elif defaults_config_file.exists():
        config.read(defaults_config_file, encoding='utf-8')

    # Überschreibe Defaults mit Werten aus Config-Datei
    for section_name, section_defaults in defaults.items():
        if section_name in config:
            for key, default_value in section_defaults.items():
                if key in config[section_name]:
                    value = config[section_name][key]
                    # Konvertiere basierend auf Default-Typ
                    # WICHTIG: bool vor int prüfen, da bool eine Subklasse von int ist
                    if isinstance(default_value, bool):
                        defaults[section_name][key] = value.lower() in ('true', 'yes', '1', 'on')
                    elif isinstance(default_value, int):
                        defaults[section_name][key] = int(value)
                    elif isinstance(default_value, float):
                        defaults[section_name][key] = float(value)
                    else:
                        defaults[section_name][key] = value

    # Lese Gantt_timeline-Defaults: hours_per_day und Schichtstart aus defaults.cfg
    _tl_hours_per_day = 8
    _tl_morning_start = '09:00'
    if defaults_config_file.exists():
        _dc = configparser.ConfigParser()
        _dc.read(defaults_config_file, encoding='utf-8')
        if _dc.has_section('WorkingHours'):
            _tl_hours_per_day = int(_dc.get('WorkingHours', 'hours_per_day', fallback='8'))
            _ms = _dc.get('WorkingHours', 'morning_shift', fallback='09:00-12:00')
            _tl_morning_start = _ms.split('-')[0].strip()
    defaults['Gantt_timeline'] = {
        'minutes_per_col': 1,
        'hours_col_minutes': 30,
        'hours_per_day': _tl_hours_per_day,
        'morning_shift_start': _tl_morning_start,
    }

    # Lese section_order aus [sections]
    _default_order = ['summary', 'critical_path', 'tasklist', 'gantt_chart', 'resource_list', 'cost_overview']
    raw_order = config.get('sections', 'section_order', fallback=None)
    defaults['sections'] = {
        'section_order': (
            [s.strip() for s in raw_order.split(',') if s.strip()]
            if raw_order else _default_order
        )
    }

    # Lese Tab-Namen aus [de]
    _default_tab_names = {
        'summary':       'Projektzusammenfassung',
        'critical_path': 'Kritischer Pfad',
        'tasklist':      'Alle Tasks',
        'netplan':       'Netzplan',
        'cost_overview': 'Kosten',
    }
    tab_names = dict(_default_tab_names)
    if config.has_section('de'):
        for key, val in config.items('de'):
            if val:
                tab_names[key] = val
    defaults['tab_names'] = tab_names

    return defaults



def is_working_day(date: datetime, working_days_str: str) -> bool:
    """
    Prüft ob ein Datum ein Arbeitstag ist basierend auf working_days Config.

    Args:
        date: Zu prüfendes Datum
        working_days_str: Kommaseparierte Liste von Wochentagen (mon,tue,wed,thu,fri,sat,sun)

    Returns:
        True wenn Arbeitstag, False sonst
    """
    weekday_map = {
        0: 'mon', 1: 'tue', 2: 'wed', 3: 'thu', 4: 'fri', 5: 'sat', 6: 'sun'
    }
    current_weekday = weekday_map[date.weekday()]
    working_days = [day.strip().lower() for day in working_days_str.split(',')]
    return current_weekday in working_days


def get_timeline_range(result: CPMResult, loadunit: str, resolution_config: Dict[str, Any]) -> Tuple[datetime, datetime, str]:
    """
    Berechnet Zeitbereich und Granularität für Timeline basierend auf minimaler Task-Dauer.

    Args:
        result: CPM-Berechnungsergebnis
        loadunit: Zeiteinheit (hours, days, etc.)
        resolution_config: Gantt_resolution Config-Werte

    Returns:
        (start_date, end_date, granularity)
        granularity kann sein: 'hours', 'halfdays', 'days'
    """
    # Berechne Projektende
    project_duration = max(task.fez for task in result.tasks.values())
    start_date = result.project_start
    end_date = add_workdays(start_date, project_duration)

    # Finde minimale Task-Dauer (ignoriere Pausen)
    min_task_duration_days = float('inf')
    for task in result.tasks.values():
        # Überspringe Pausen-Tasks
        is_break = hasattr(task, 'is_break') and task.is_break
        if not is_break and task.duration > 0:
            min_task_duration_days = min(min_task_duration_days, task.duration)

    # Wenn keine regulären Tasks gefunden wurden, verwende Standardwert
    if min_task_duration_days == float('inf'):
        min_task_duration_days = 1.0

    # Prüfe zuerst result.time_unit für direkte Zuordnung
    if hasattr(result, 'time_unit'):
        if result.time_unit == 'minutes':
            # Minuten-Projekt → Minuten-Granularität
            granularity = 'minutes'
            return (start_date, end_date, granularity)
        elif result.time_unit == 'hours':
            # Stunden-Projekt → Stunden-Granularität
            granularity = 'hours'
            return (start_date, end_date, granularity)
        # Für 'days' verwenden wir die normale Logik unten

    # Konvertiere minimale Dauer in Minuten (8h Arbeitstag = 1 Tag)
    min_task_duration_min = min_task_duration_days * 8 * 60

    # Bestimme Granularität basierend auf Config-Schwellenwerten (Fallback)
    threshold_hours = resolution_config['threshold_hours_min']
    threshold_halfday = resolution_config['threshold_halfday_min']

    if min_task_duration_min < threshold_hours:
        granularity = 'hours'
    elif min_task_duration_min < threshold_halfday:
        granularity = 'halfdays'
    else:
        granularity = 'days'

    # Überschreibe mit loadunit falls angegeben
    if loadunit == 'hours':
        granularity = 'hours'
    elif loadunit == 'days':
        granularity = 'days'

    return start_date, end_date, granularity


def create_timeline_header(ws, start_col: int, start_row: int, start_date: datetime,
                          end_date: datetime, granularity: str,
                          working_days: str = 'mon,tue,wed,thu,fri',
                          timeline_config: dict = None) -> tuple:
    """
    Erstellt Timeline-Header (Zeitstrahl) in Excel.

    Args:
        ws: Worksheet
        start_col: Start-Spalte für Timeline
        start_row: Start-Zeile für Header
        start_date: Projekt-Startdatum
        end_date: Projekt-Enddatum
        granularity: 'minutes', 'hours', 'halfdays', 'days'
        working_days: Kommaseparierte Liste von Arbeitstagen (z.B. 'mon,tue,wed,thu,fri')
        timeline_config: Gantt_timeline-Konfiguration (aus excel_export.cfg)

    Returns:
        (rows_used, cols_per_workday): Header-Zeilenanzahl und Spalten pro Arbeitstag
    """
    tlcfg = timeline_config or {}
    duration_days = (end_date - start_date).days

    # Header-Stil
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=9)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_col = start_col
    rows_used = 1

    if granularity == 'minutes':
        # Minutengranularität: jede Spalte = minutes_per_col Minuten
        minutes_per_col   = int(tlcfg.get('minutes_per_col', 1))
        hours_per_day     = int(tlcfg.get('hours_per_day', 8))
        morning_start_str = tlcfg.get('morning_shift_start', '09:00')
        mh, mm_s          = map(int, morning_start_str.split(':'))
        morning_start_min = mh * 60 + mm_s
        cols_per_workday  = (hours_per_day * 60) // minutes_per_col
        col_width         = max(1.2, 12.0 / max(1, cols_per_workday // 8))

        rows_used     = 2  # Datum + Uhrzeit
        weekend_fill  = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        weekday_names = {0: 'Mo', 1: 'Di', 2: 'Mi', 3: 'Do', 4: 'Fr', 5: 'Sa', 6: 'So'}
        label_every   = max(1, 60 // minutes_per_col)  # Beschriftung jede volle Stunde

        current_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        while current_date.date() <= end_date.date():
            is_workday = is_working_day(current_date, working_days)
            if is_workday:
                # Zeile 1: Datum über alle Minuten-Spalten
                ws.merge_cells(
                    start_row=start_row, start_column=current_col,
                    end_row=start_row,   end_column=current_col + cols_per_workday - 1
                )
                cell = ws.cell(row=start_row, column=current_col,
                               value=current_date.strftime('%d.%m'))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

                # Zeile 2: HH:MM jede volle Stunde, sonst leer
                for col_i in range(cols_per_workday):
                    abs_min = morning_start_min + col_i * minutes_per_col
                    h, m = divmod(abs_min, 60)
                    label = f'{h:02d}:{m:02d}' if col_i % label_every == 0 else ''
                    cell = ws.cell(row=start_row + 1, column=current_col + col_i, value=label)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                    ws.column_dimensions[get_column_letter(current_col + col_i)].width = col_width

                current_col += cols_per_workday
            else:
                # Nicht-Arbeitstag: 1 Spalte
                cell = ws.cell(row=start_row, column=current_col,
                               value=current_date.strftime('%d.%m'))
                cell.fill = weekend_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                weekday_label = weekday_names[current_date.weekday()]
                cell2 = ws.cell(row=start_row + 1, column=current_col, value=weekday_label)
                cell2.fill = weekend_fill
                cell2.font = header_font
                cell2.alignment = header_alignment
                cell2.border = border
                ws.column_dimensions[get_column_letter(current_col)].width = col_width
                current_col += 1

            current_date += timedelta(days=1)

    elif granularity == 'hours':
        # Stundengranularität: jede Spalte = hours_col_minutes Minuten (Standard: 30 = halbe Stunde)
        hours_col_minutes = int(tlcfg.get('hours_col_minutes', 30))
        hours_per_day     = int(tlcfg.get('hours_per_day', 8))
        morning_start_str = tlcfg.get('morning_shift_start', '09:00')
        mh, mm_s          = map(int, morning_start_str.split(':'))
        morning_start_min = mh * 60 + mm_s
        cols_per_workday  = (hours_per_day * 60) // hours_col_minutes

        rows_used     = 2  # Datum + Uhrzeit
        weekend_fill  = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        weekday_names = {0: 'Mo', 1: 'Di', 2: 'Mi', 3: 'Do', 4: 'Fr', 5: 'Sa', 6: 'So'}

        current_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        while current_date.date() <= end_date.date():
            is_workday = is_working_day(current_date, working_days)
            if is_workday:
                # Zeile 1: Datum über alle Slot-Spalten
                ws.merge_cells(
                    start_row=start_row, start_column=current_col,
                    end_row=start_row,   end_column=current_col + cols_per_workday - 1
                )
                cell = ws.cell(row=start_row, column=current_col,
                               value=current_date.strftime('%d.%m'))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

                # Zeile 2: HH:MM für jeden Slot
                for slot_i in range(cols_per_workday):
                    abs_min = morning_start_min + slot_i * hours_col_minutes
                    h, m = divmod(abs_min, 60)
                    cell = ws.cell(row=start_row + 1, column=current_col + slot_i,
                                   value=f'{h:02d}:{m:02d}')
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                    ws.column_dimensions[get_column_letter(current_col + slot_i)].width = 5

                current_col += cols_per_workday
            else:
                # Nicht-Arbeitstag: 1 Spalte
                cell = ws.cell(row=start_row, column=current_col,
                               value=current_date.strftime('%d.%m'))
                cell.fill = weekend_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                weekday_label = weekday_names[current_date.weekday()]
                cell2 = ws.cell(row=start_row + 1, column=current_col, value=weekday_label)
                cell2.fill = weekend_fill
                cell2.font = header_font
                cell2.alignment = header_alignment
                cell2.border = border
                ws.column_dimensions[get_column_letter(current_col)].width = 4
                current_col += 1

            current_date += timedelta(days=1)

    elif granularity == 'halfdays':
        # Halbtages-basierte Timeline (Vormittag/Nachmittag)
        cols_per_workday = 2
        # Arbeitstage: 2 Spalten (VM/NM)
        # Nicht-Arbeitstage: 1 Spalte pro Tag (Sa, So)
        rows_used = 2  # Datum und VM/NM/Sa/So

        # Stilvorlagen
        weekend_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        weekday_names = {0: 'Mo', 1: 'Di', 2: 'Mi', 3: 'Do', 4: 'Fr', 5: 'Sa', 6: 'So'}

        # Erste Pass: Berechne Spalten pro Datum und erstelle Mapping
        date_col_mapping = {}  # Datum -> (start_col, num_cols, is_workday)
        temp_date = start_date
        temp_col = start_col

        while temp_date <= end_date:
            is_workday = is_working_day(temp_date, working_days)
            num_cols = 2 if is_workday else 1  # Arbeitstag = 2 Spalten, sonst 1
            date_col_mapping[temp_date.date()] = (temp_col, num_cols, is_workday)
            temp_col += num_cols
            temp_date += timedelta(days=1)

        # Zeile 1: Datum-Header mit Merging
        current_date = start_date
        while current_date <= end_date:
            start_col_for_date, num_cols, is_workday = date_col_mapping[current_date.date()]

            # Merge Spalten wenn mehrere
            if num_cols > 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=start_col_for_date,
                    end_row=start_row,
                    end_column=start_col_for_date + num_cols - 1
                )

            cell = ws.cell(row=start_row, column=start_col_for_date,
                         value=current_date.strftime('%d.%m'))
            cell.fill = weekend_fill if not is_workday else header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

            current_date += timedelta(days=1)

        # Zeile 2: VM/NM für Arbeitstage, Wochentag-Kürzel für Nicht-Arbeitstage
        current_date = start_date

        while current_date <= end_date:
            start_col_for_date, num_cols, is_workday = date_col_mapping[current_date.date()]

            if is_workday:
                # Arbeitstag: Vormittag und Nachmittag
                # Vormittag
                cell = ws.cell(row=start_row + 1, column=start_col_for_date, value='VM')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(start_col_for_date)].width = 5

                # Nachmittag
                cell = ws.cell(row=start_row + 1, column=start_col_for_date + 1, value='NM')
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(start_col_for_date + 1)].width = 5
            else:
                # Nicht-Arbeitstag: Wochentag-Kürzel (Sa, So, etc.)
                weekday_label = weekday_names[current_date.weekday()]
                cell = ws.cell(row=start_row + 1, column=start_col_for_date, value=weekday_label)
                cell.fill = weekend_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(start_col_for_date)].width = 5

            current_date += timedelta(days=1)

    elif granularity == 'days':
        # Tagesgranularität: 1 Spalte pro Tag
        cols_per_workday = 1
        # Wenn mehr als 7 Tage: Zeige Kalenderwochen
        # Wochenenden-Hintergrund
        weekend_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        if duration_days > 7:
            rows_used = 3  # Monat, Tag, KW

            # Zeile 1: Monat
            current_date = start_date
            month_start_col = current_col
            current_month = current_date.month
            month_days = 0

            while current_date <= end_date:
                if current_date.month != current_month:
                    # Merge Monatszellen
                    if month_days > 1:
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=month_start_col,
                            end_row=start_row,
                            end_column=month_start_col + month_days - 1
                        )
                    cell = ws.cell(row=start_row, column=month_start_col,
                                 value=start_date.replace(day=1).strftime('%B %Y'))
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border

                    month_start_col = current_col
                    current_month = current_date.month
                    month_days = 0

                month_days += 1
                current_date += timedelta(days=1)
                current_col += 1

            # Letzter Monat
            if month_days > 0:
                if month_days > 1:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=month_start_col,
                        end_row=start_row,
                        end_column=month_start_col + month_days - 1
                    )
                last_month_date = end_date.replace(day=1)
                cell = ws.cell(row=start_row, column=month_start_col,
                             value=last_month_date.strftime('%B %Y'))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Zeile 2: Tage (mit Wochenendmarkierung)
            current_date = start_date
            current_col = start_col
            while current_date <= end_date:
                is_workday = is_working_day(current_date, working_days)
                cell = ws.cell(row=start_row + 1, column=current_col, value=current_date.day)
                cell.fill = header_fill if is_workday else weekend_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(current_col)].width = 4

                current_date += timedelta(days=1)
                current_col += 1

            # Zeile 3: Kalenderwochen (mit Wochenendmarkierung)
            current_date = start_date
            current_col = start_col
            cw_start_col = current_col
            current_cw = current_date.isocalendar()[1]
            cw_days = 0

            while current_date <= end_date:
                date_cw = current_date.isocalendar()[1]

                if date_cw != current_cw:
                    # Merge KW-Zellen
                    if cw_days > 1:
                        ws.merge_cells(
                            start_row=start_row + 2,
                            start_column=cw_start_col,
                            end_row=start_row + 2,
                            end_column=cw_start_col + cw_days - 1
                        )
                    cell = ws.cell(row=start_row + 2, column=cw_start_col, value=f"KW{current_cw}")
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border

                    cw_start_col = current_col
                    current_cw = date_cw
                    cw_days = 0

                cw_days += 1
                current_date += timedelta(days=1)
                current_col += 1

            # Letzte KW
            if cw_days > 0:
                if cw_days > 1:
                    ws.merge_cells(
                        start_row=start_row + 2,
                        start_column=cw_start_col,
                        end_row=start_row + 2,
                        end_column=cw_start_col + cw_days - 1
                    )
                cell = ws.cell(row=start_row + 2, column=cw_start_col, value=f"KW{current_cw}")
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

        else:
            # Nur Tage (ohne KW) - mit Wochenendmarkierung
            current_date = start_date
            while current_date <= end_date:
                is_workday = is_working_day(current_date, working_days)
                cell = ws.cell(row=start_row, column=current_col,
                             value=current_date.strftime('%d.%m'))
                cell.fill = header_fill if is_workday else weekend_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(current_col)].width = 8

                current_date += timedelta(days=1)
                current_col += 1

    return rows_used, cols_per_workday


def draw_gantt_bar(ws, row: int, start_col: int, faz_offset: int, duration_cols: int,
                  sez_offset: int, color: str, critical: bool, config: Dict[str, Any]) -> None:
    """
    Zeichnet Gantt-Balken mit zwei Rechtecken (transparent für Puffer, gefüllt für Dauer).

    Args:
        ws: Worksheet
        row: Zeile für den Balken
        start_col: Start-Spalte des Chart-Bereichs
        faz_offset: Offset für FAZ (Frühester Anfangszeitpunkt)
        duration_cols: Anzahl Spalten für Dauer
        sez_offset: Offset für SEZ (Spätester Endzeitpunkt)
        color: Balken-Farbe (Hex)
        critical: Ist Task auf kritischem Pfad?
        config: Gantt-Konfiguration
    """
    border_color = config['critical_border_color'] if critical else None

    # Transparenter Balken von FAZ bis SEZ (Puffer)
    for col_offset in range(faz_offset, sez_offset):
        cell = ws.cell(row=row, column=start_col + col_offset)
        # Transparente Füllung (hellere Version der Farbe)
        # Stelle sicher, dass color 6 Zeichen hat (füge 00 prefix hinzu falls nötig)
        color_normalized = color if len(color) >= 6 else "00" + color
        cell.fill = PatternFill(start_color=color_normalized + '40', end_color=color_normalized + '40', fill_type="solid")

        if critical and border_color:
            # Rote Umrandung für kritischen Pfad
            border_side = Side(style='medium', color=border_color)
            if col_offset == faz_offset:
                cell.border = Border(left=border_side, top=border_side, bottom=border_side)
            elif col_offset == sez_offset - 1:
                cell.border = Border(right=border_side, top=border_side, bottom=border_side)
            else:
                cell.border = Border(top=border_side, bottom=border_side)

    # Gefüllter Balken von FAZ bis FAZ+Dauer
    for col_offset in range(faz_offset, faz_offset + duration_cols):
        cell = ws.cell(row=row, column=start_col + col_offset)
        # Normalisiere Farbe (füge 00 prefix hinzu falls nötig)
        color_normalized = color if len(color) >= 6 else "00" + color
        cell.fill = PatternFill(start_color=color_normalized, end_color=color_normalized, fill_type="solid")

        if critical and border_color:
            # Rote Umrandung für kritischen Pfad
            border_side = Side(style='medium', color=border_color)
            if col_offset == faz_offset:
                cell.border = Border(left=border_side, top=border_side, bottom=border_side)
            elif col_offset == faz_offset + duration_cols - 1:
                cell.border = Border(right=border_side, top=border_side, bottom=border_side)
            else:
                cell.border = Border(top=border_side, bottom=border_side)


def create_gantt_chart(wb: Workbook, project: Project, result: CPMResult,
                      report: GanttReport, full_config: Dict[str, Any]) -> None:
    """
    Erstellt Gantt-Chart-Worksheet.

    Args:
        wb: Workbook
        project: Projekt-Daten
        result: CPM-Berechnungsergebnis
        report: Gantt Report-Konfiguration
        full_config: Vollständige Excel-Export-Konfiguration (alle Sektionen)
    """
    # Extrahiere die relevanten Config-Sektionen
    config = full_config['GanttChart']
    resolution_config = full_config['Gantt_resolution']

    ws = wb.create_sheet(title=report.name)

    # Headline
    ws['A1'] = report.headline
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    # Spalten-Header
    row = 3
    for col_idx, col_name in enumerate(report.columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # Setze Spaltenbreiten aus Config
    ws.column_dimensions['A'].width = config['col_width_id']
    ws.column_dimensions['B'].width = config['col_width_name']
    ws.column_dimensions['C'].width = config['col_width_start']
    ws.column_dimensions['D'].width = config['col_width_end']
    ws.column_dimensions['E'].width = config['col_width_effort']

    # Timeline berechnen mit Resolution-Config
    start_date, end_date, granularity = get_timeline_range(result, report.loadunit, resolution_config)

    # Timeline-Header (ab Spalte 6 = 'chart')
    chart_start_col = len(report.columns)
    working_days = resolution_config['working_days']
    timeline_config = full_config.get('Gantt_timeline', {})
    timeline_rows, cols_per_workday = create_timeline_header(
        ws, chart_start_col, row, start_date, end_date, granularity, working_days, timeline_config
    )

    # Daten-Zeilen
    data_start_row = row + timeline_rows
    current_row = data_start_row

    # Sortiere Tasks nach FAZ (Frühester Anfangszeitpunkt), dann nach ID
    sorted_task_ids = sorted(
        result.tasks.keys(),
        key=lambda x: (result.tasks[x].faz, str(x))
    )

    # Kritischer Pfad
    critical_path = result.critical_path

    # Farb-Interpolation
    num_tasks = len(sorted_task_ids)

    for task_idx, task_id in enumerate(sorted_task_ids):
        task = result.tasks[task_id]

        # Prüfe ob Task eine Pause ist
        is_break = hasattr(task, 'is_break') and task.is_break

        # Interpoliere Farbe (Pausen werden grau)
        if is_break:
            bar_color = "D3D3D3"  # Grau für Pausen
        else:
            if num_tasks > 1:
                color_position = task_idx / (num_tasks - 1)
            else:
                color_position = 0.0
            bar_color = interpolate_color(config['color_start'], config['color_end'], color_position)

        # Spalte A: Vorgang (Task-ID)
        ws.cell(row=current_row, column=1, value=str(task_id))

        # Spalte B: name (Task-Name, nicht Person!)
        ws.cell(row=current_row, column=2, value=task.name)

        # Spalte C: start (leer, da nicht in Task-Daten)
        ws.cell(row=current_row, column=3, value='')

        # Spalte D: end (leer, da nicht in Task-Daten)
        ws.cell(row=current_row, column=4, value='')

        # Spalte E: effort
        ws.cell(row=current_row, column=5, value=format_time_value_auto(task.duration))

        # Spalte F+: chart (Balkendiagramm)
        # cols_per_workday bestimmt die Skalierung: task.faz/duration in Tagen × cols_per_workday = Spalten-Offset
        faz_offset    = int(task.faz      * cols_per_workday)
        duration_cols = max(1, int(task.duration * cols_per_workday))
        sez_offset    = int(task.sez      * cols_per_workday)

        is_critical = task_id in critical_path

        draw_gantt_bar(
            ws, current_row, chart_start_col, faz_offset, duration_cols, sez_offset,
            bar_color, is_critical, config
        )

        current_row += 1


def create_resource_list(wb: Workbook, project: Project, result: CPMResult,
                        report: ResourceListReport, full_config: Dict[str, Any]) -> None:
    """
    Erstellt Resource-List-Worksheet.

    Args:
        wb: Workbook
        project: Projekt-Daten
        result: CPM-Berechnungsergebnis
        report: Resource List Report-Konfiguration
        full_config: Vollständige Excel-Export-Konfiguration (alle Sektionen)
    """
    # Extrahiere die relevanten Config-Sektionen
    config = full_config['GanttChart']  # Für Farben
    resource_config = full_config['ResourceList']
    resolution_config = full_config['Gantt_resolution']

    ws = wb.create_sheet(title=report.name)

    # Headline
    ws['A1'] = report.headline
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # Spalten-Header
    row = 3
    for col_idx, col_name in enumerate(report.columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # Setze Spaltenbreiten aus Config
    ws.column_dimensions['A'].width = resource_config['col_width_resource']
    ws.column_dimensions['B'].width = resource_config['col_width_task']
    ws.column_dimensions['C'].width = resource_config['col_width_start']
    ws.column_dimensions['D'].width = resource_config['col_width_end']

    # Timeline berechnen mit Resolution-Config
    start_date, end_date, granularity = get_timeline_range(result, report.loadunit, resolution_config)

    # Timeline-Header (ab Spalte 5 = 'chart')
    chart_start_col = len(report.columns)
    working_days = resolution_config['working_days']
    timeline_config = full_config.get('Gantt_timeline', {})
    timeline_rows, cols_per_workday = create_timeline_header(
        ws, chart_start_col, row, start_date, end_date, granularity, working_days, timeline_config
    )

    # Daten-Zeilen (Personen und Maschinen)
    data_start_row = row + timeline_rows
    current_row = data_start_row

    # Erstelle Ressource -> Tasks Mapping
    resource_tasks = {}
    for res in project.resources:
        resource_tasks[res.id] = []

    # Sammle Tasks pro Ressource
    for task_id in result.tasks.keys():
        task = next((t for t in project.tasks if t.id == task_id), None)
        if task and task.resources:
            for res_id in task.resources:
                if res_id in resource_tasks:
                    resource_tasks[res_id].append(task_id)

    # Erstelle Liste aller anzuzeigenden Ressourcen (Personen + Maschinen)
    display_resources = []

    # Erst Personen
    for person in project.persons:
        person_res = next((r for r in project.resources if hasattr(r, 'person_id') and r.person_id == person.id), None)
        if person_res:
            display_resources.append({
                'type': 'person',
                'id': person_res.id,
                'name': person.name,
                'role': person.role if hasattr(person, 'role') else '',
                'person_id': person.id,
                'color': person_res.color if hasattr(person_res, 'color') else None
            })

    # Dann Maschinen (machine, truck)
    for res in project.resources:
        if res.type in ['machine', 'truck']:
            display_resources.append({
                'type': res.type,
                'id': res.id,
                'name': res.name,
                'role': res.type.capitalize(),
                'person_id': None,
                'color': res.color if hasattr(res, 'color') else None
            })

    # Farb-Zuweisung für alle Ressourcen
    num_resources = len(display_resources)
    resource_colors = {}
    for res_idx, res_info in enumerate(display_resources):
        # Verwende benutzerdefinierte Farbe falls vorhanden, sonst interpolierte Farbe
        if res_info.get('color'):
            resource_colors[res_info['id']] = res_info['color']
        else:
            # Fallback: Farb-Interpolation
            if num_resources > 1:
                color_position = res_idx / (num_resources - 1)
            else:
                color_position = 0.0
            resource_colors[res_info['id']] = interpolate_color(config['color_start'], config['color_end'], color_position)

    # cols_per_workday bestimmt die Skalierung (konsistent mit Gantt-Chart)

    # Zeige alle Ressourcen
    for res_info in display_resources:
        res_color = resource_colors[res_info['id']]

        # Spalte A: Name mit ID (mit Ressourcenfarbe als Hintergrund)
        cell_name = ws.cell(row=current_row, column=1, value=f"{res_info['name']} ({res_info['id']})")
        # Füge leichte Hintergrundfarbe hinzu
        color_with_transparency = res_color if len(res_color) >= 6 else f"00{res_color}"
        cell_name.fill = PatternFill(start_color=color_with_transparency + "40", end_color=color_with_transparency + "40", fill_type="solid")

        # Spalte B: Rolle/Typ
        cell_role = ws.cell(row=current_row, column=2, value=res_info['role'])
        cell_role.fill = PatternFill(start_color=color_with_transparency + "40", end_color=color_with_transparency + "40", fill_type="solid")

        # Spalte C: start (leer)
        ws.cell(row=current_row, column=3, value='')

        # Spalte D: end (leer)
        ws.cell(row=current_row, column=4, value='')

        # Spalte E+: chart (Balken für Tasks)
        for task_id in resource_tasks[res_info['id']]:
            task = result.tasks[task_id]

            faz_offset = int(task.faz * cols_per_workday)
            duration_cols = max(1, int(task.duration * cols_per_workday))

            # Prüfe ob Task eine Pause ist
            is_break = hasattr(task, 'is_break') and task.is_break

            # Zeichne Balken (ohne Puffer, nur Dauer)
            for col_offset in range(faz_offset, faz_offset + duration_cols):
                cell = ws.cell(row=current_row, column=chart_start_col + col_offset)
                if is_break:
                    # Pausen werden in Grau dargestellt
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                else:
                    # Normale Arbeitszeit in Ressourcenfarbe
                    cell.fill = PatternFill(start_color=res_color, end_color=res_color, fill_type="solid")

        current_row += 1

    # Färbe Wochenend-Spalten grau (nur leere Zellen)
    if granularity == 'days':
        weekend_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        duration_days = (end_date - start_date).days + 1

        for day_offset in range(duration_days):
            current_date = start_date + timedelta(days=day_offset)
            # Prüfe ob Wochenende (Samstag=5, Sonntag=6)
            if current_date.weekday() in [5, 6]:
                weekend_col = chart_start_col + day_offset
                # Färbe diese Spalte für alle Ressourcen-Zeilen (nur wenn leer)
                for row_offset in range(len(display_resources)):
                    cell = ws.cell(row=data_start_row + row_offset, column=weekend_col)
                    # Nur färben wenn Zelle noch keine Füllung hat (leer)
                    if not cell.fill or cell.fill.start_color.index in ['00000000', 'FFFFFFFF', None]:
                        cell.fill = weekend_fill


def _apply_header_style(cell, fill_color: str = "4472C4") -> None:
    """Wendet Standard-Header-Stil auf eine Zelle an."""
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center')


def create_summary_sheet(wb: Workbook, result: CPMResult, project_name: str,
                         tab_name: str = "Projektzusammenfassung") -> None:
    """
    Erzeugt einen Zusammenfassungs-Tab mit Projektübersicht.

    Enthält: Projektname, Dauer, Start, Ende, Zeiteinheit,
             Anzahl Tasks, Anzahl kritische Tasks, Puffer-Statistik.

    Args:
        wb:           Workbook
        result:       CPM-Berechnungsergebnis
        project_name: Projektname
        tab_name:     Reiter-Bezeichnung (aus [de]-Sektion der CFG)
    """
    ws = wb.create_sheet(title=tab_name)

    label_font = Font(bold=True)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)

    # Titel
    ws['A1'] = "Projektzusammenfassung"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 28

    # Projektdaten
    project_duration = max(task.fez for task in result.tasks.values())
    critical_tasks = [t for t in result.tasks.values() if t.is_critical]
    non_break_tasks = [t for t in result.tasks.values()
                       if not t.is_break and not (isinstance(t.id, str) and t.id.startswith('WE-'))]

    rows = [
        ("Projekt:",         project_name),
        ("Projektdauer:",    format_time_value_auto(project_duration)),
        ("Zeiteinheit:",     result.time_unit),
        ("Anzahl Tasks:",    len(non_break_tasks)),
        ("Kritische Tasks:", len([t for t in non_break_tasks if t.is_critical])),
    ]

    if result.project_start:
        end_date = add_workdays(result.project_start, project_duration)
        rows.insert(2, ("Startdatum:", result.project_start.strftime('%Y-%m-%d')))
        rows.insert(3, ("Enddatum:",   end_date.strftime('%Y-%m-%d')))

    if result.calculation_date:
        rows.append(("Erstellt am:", result.calculation_date.strftime('%Y-%m-%d %H:%M')))

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value)

    # Puffer-Statistik
    puffer_row = len(rows) + 4
    ws.cell(row=puffer_row, column=1, value="Puffer-Statistik").font = Font(bold=True, size=11)
    puffer_row += 1

    headers = ["Task", "Dauer", "GP", "FP", "Kritisch"]
    for col, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=puffer_row, column=col))
        ws.cell(row=puffer_row, column=col).value = h

    puffer_row += 1
    crit_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    for task in sorted(non_break_tasks, key=lambda t: t.faz):
        ws.cell(row=puffer_row, column=1, value=task.name)
        ws.cell(row=puffer_row, column=2, value=format_time_value_auto(task.duration))
        ws.cell(row=puffer_row, column=3, value=format_time_value_auto(task.puffer))
        ws.cell(row=puffer_row, column=4, value=format_time_value_auto(task.free_puffer))
        ws.cell(row=puffer_row, column=5, value="JA" if task.is_critical else "")
        if task.is_critical:
            for col in range(1, 6):
                ws.cell(row=puffer_row, column=col).fill = crit_fill
        puffer_row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10


def create_critical_path_sheet(wb: Workbook, result: CPMResult,
                               tab_name: str = "Kritischer Pfad") -> None:
    """
    Erzeugt einen Tab mit ausschließlich den kritischen Tasks.

    Enthält: ID, Name, Dauer, FAZ, FEZ (in der Zeiteinheit des Projekts).

    Args:
        wb:       Workbook
        result:   CPM-Berechnungsergebnis
        tab_name: Reiter-Bezeichnung (aus [de]-Sektion der CFG)
    """
    ws = wb.create_sheet(title=tab_name)

    title_fill = PatternFill(start_color="C92A2A", end_color="C92A2A", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    crit_fill  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    ws['A1'] = "Kritischer Pfad"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 28

    tu = result.time_unit
    headers = ["ID", "Name", f"Dauer ({tu})", f"FAZ ({tu})", f"FEZ ({tu})", f"GP ({tu})"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        _apply_header_style(cell, fill_color="C92A2A")

    row = 4
    critical_ids = [tid for tid in result.critical_path
                    if tid in result.tasks
                    and not (isinstance(tid, str) and tid.startswith('WE-'))
                    and not result.tasks[tid].is_break]

    for tid in critical_ids:
        task = result.tasks[tid]
        ws.cell(row=row, column=1, value=str(tid))
        ws.cell(row=row, column=2, value=task.name)
        ws.cell(row=row, column=3, value=format_time_value_auto(task.duration))
        ws.cell(row=row, column=4, value=format_time_value_auto(task.faz))
        ws.cell(row=row, column=5, value=format_time_value_auto(task.fez))
        ws.cell(row=row, column=6, value=format_time_value_auto(task.puffer))
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = crit_fill
        row += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12


def create_tasklist_sheet(wb: Workbook, result: CPMResult,
                          tab_name: str = "Alle Tasks") -> None:
    """
    Erzeugt einen vollständigen Tasklisten-Tab mit allen CPM-Werten.

    Entspricht dem früheren "CPM Analyse"-Tab: ID, Name, Dauer,
    FAZ, FEZ, SAZ, SEZ, Puffer, Kritisch.

    Args:
        wb:       Workbook
        result:   CPM-Berechnungsergebnis
        tab_name: Reiter-Bezeichnung (aus [de]-Sektion der CFG)
    """
    ws = wb.create_sheet(title=tab_name)

    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    crit_fill  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    ws['A1'] = "Taskliste"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:I1')
    ws.row_dimensions[1].height = 28

    tu = result.time_unit
    headers = ['ID', 'Name', 'Dauer', f'FAZ ({tu})', f'FEZ ({tu})',
               f'SAZ ({tu})', f'SEZ ({tu})', 'Puffer', 'Kritisch']
    for col, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=3, column=col, value=h))

    sorted_ids = sorted(
        [tid for tid in result.tasks
         if not (isinstance(tid, str) and tid.startswith('WE-'))
         and not result.tasks[tid].is_break],
        key=lambda x: result.tasks[x].faz
    )

    row = 4
    for tid in sorted_ids:
        task = result.tasks[tid]
        ws.cell(row=row, column=1, value=str(tid))
        ws.cell(row=row, column=2, value=task.name)
        ws.cell(row=row, column=3, value=format_time_value_auto(task.duration))
        ws.cell(row=row, column=4, value=format_time_value_auto(task.faz))
        ws.cell(row=row, column=5, value=format_time_value_auto(task.fez))
        ws.cell(row=row, column=6, value=format_time_value_auto(task.saz))
        ws.cell(row=row, column=7, value=format_time_value_auto(task.sez))
        ws.cell(row=row, column=8, value=format_time_value_auto(task.puffer))
        ws.cell(row=row, column=9, value="JA" if task.is_critical else "")
        if task.is_critical:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = crit_fill
        row += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10


def create_netplan_table_sheet(wb: Workbook, result: CPMResult, tab_name: str = "Netzplan") -> None:
    """
    Erzeugt einen Tab mit Netzplan-Daten als Tabelle (aus generate_network_csv).

    Args:
        wb:       Workbook
        result:   CPM-Berechnungsergebnis
        tab_name: Reiter-Bezeichnung
    """
    csv_text = result.to_network_csv()
    reader = csv.reader(io.StringIO(csv_text))
    rows = list(reader)

    ws = wb.create_sheet(title=tab_name)

    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    crit_fill  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    ws['A1'] = tab_name
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    if rows:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
    ws.row_dimensions[1].height = 28

    for row_idx, cols in enumerate(rows, start=3):
        is_header = row_idx == 3
        is_critical = len(cols) >= 11 and cols[10].strip() == 'JA'

        for col_idx, value in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_header:
                _apply_header_style(cell)
            elif is_critical:
                cell.fill = crit_fill

    col_widths = [10, 35, 8, 8, 8, 8, 8, 8, 8, 25, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_cost_sheet(wb: 'Workbook', project: Any, result: CPMResult,
                      tab_name: str = "Kosten") -> None:
    """
    Erzeugt einen Kosten-Tab mit Personalkosten, Maschinenkosten und Bereitstellungskosten.

    Args:
        wb:       Workbook
        project:  Projekt-Objekt mit persons und resources
        result:   CPM-Berechnungsergebnis
        tab_name: Reiter-Bezeichnung
    """
    try:
        from cost_calculator import calculate_project_costs
    except ImportError:
        from .cost_calculator import calculate_project_costs

    ws = wb.create_sheet(title=tab_name)

    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subtotal_fill = PatternFill(start_color="EEF2FA", end_color="EEF2FA", fill_type="solid")

    ws['A1'] = tab_name
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 28

    costs = calculate_project_costs(project, result)

    if not costs:
        ws['A3'] = "Keine Kostendaten verfügbar – Stundensätze fehlen."
        ws.column_dimensions['A'].width = 50
        return

    # Spaltenköpfe
    headers = ['Ressource', 'Typ', 'Stunden', '€/h', 'Bereitst. €', 'Lohnkosten €', 'Gesamt €']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        _apply_header_style(cell)

    person_fill  = PatternFill(start_color="EAF3FB", end_color="EAF3FB", fill_type="solid")
    machine_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    money_align  = Alignment(horizontal='right')

    row = 4
    for e in costs.entries:
        row_fill = person_fill if e.resource_type == 'person' else machine_fill
        values = [
            e.resource_name,
            e.resource_type,
            e.hours_worked,
            e.hourly_rate,
            e.provisioning_costs,
            e.labor_costs,
            e.total_costs,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            if col >= 3:
                cell.alignment = money_align
                if col >= 4 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00 "€"'
        row += 1

    # Leerzeile
    row += 1

    # Summenzeilen
    def _sum_row(label: str, value: float, fill: 'PatternFill') -> None:
        nonlocal row
        ws.cell(row=row, column=5, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=7, value=value)
        cell.number_format = '#,##0.00 "€"'
        cell.font = Font(bold=True)
        cell.alignment = money_align
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill
        row += 1

    if costs.total_person_costs > 0:
        _sum_row("Personalkosten", costs.total_person_costs, subtotal_fill)
    if costs.total_machine_costs > 0:
        _sum_row("Maschinenkosten (Lohn)", costs.total_machine_costs, subtotal_fill)
    if costs.total_provisioning_costs > 0:
        _sum_row("Bereitstellungskosten", costs.total_provisioning_costs, subtotal_fill)

    row += 1
    _sum_row("GESAMTKOSTEN", costs.total_costs, total_fill)

    # Spaltenbreiten
    col_widths = [30, 12, 10, 10, 14, 14, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

