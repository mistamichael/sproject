#!/usr/bin/env python3
"""
Comprehensive unit tests for sproject
======================================

Tests all major components:
- Pydantic models (unified Project class with aliases)
- CPM calculation
- Report models
- Excel export
- TXT export (inkl. netplan_table)
- Project expansion (cycles, loops)
- Network diagram (generate_network_csv)
"""

import os
import unittest
import sys
import tempfile
from pathlib import Path
from datetime import datetime

# Add lib to path
sys.path.insert(0, str(Path(os.environ.get("PV_LIB", str(Path(__file__).parent.parent / "lib")))))
sys.path.insert(0, str(Path(os.environ.get("PROJECT", str(Path(__file__).parent.parent)))))

_EXAMPLES_DIR = Path(os.environ.get("PV_EXAMPLES", str(Path(__file__).parent.parent / "examples")))
_CFG_DIR      = Path(os.environ.get("PV_CFG",      str(Path(__file__).parent.parent / "cfg")))

from lib.models import (
    load_project, save_project,
    Project, SimpleProject, PersonProject, CycleProject, LoopProject,
    CPMCalculator,
    InstanceTask, LoopTask,
)


class TestProjectLoading(unittest.TestCase):
    """Tests für das Laden verschiedener Projekt-Typen."""

    def test_load_simple_project(self):
        """Test: Einfaches Projekt laden (tankdesign.json)"""
        tankdesign_file = _EXAMPLES_DIR / 'tankdesign.json'

        if not tankdesign_file.exists():
            self.skipTest(f"Datei nicht gefunden: {tankdesign_file}")

        project = load_project(tankdesign_file)

        self.assertIsInstance(project, Project)
        self.assertIsNotNone(project.project)
        self.assertGreater(len(project.tasks), 0)

        first_task = project.tasks[0]
        self.assertIsNotNone(first_task.id)
        self.assertIsNotNone(first_task.name)

    def test_load_person_project(self):
        """Test: Projekt mit Personen laden (software_simple.json)"""
        software_file = _EXAMPLES_DIR / 'software_simple.json'
        if not software_file.exists():
            self.skipTest(f"Datei nicht gefunden: {software_file}")

        project = load_project(software_file)

        self.assertIsInstance(project, Project)
        self.assertEqual(project.project, "Software Entwicklung Projekt")
        self.assertIsNotNone(project.persons)
        self.assertGreater(len(project.persons), 0)
        self.assertIsNotNone(project.resources)
        self.assertGreater(len(project.resources), 0)
        self.assertGreater(len(project.tasks), 0)

    def test_load_cycle_project(self):
        """Test: Projekt mit InstanceTasks laden (pizzas.json)"""
        pizzas_file = _EXAMPLES_DIR / 'pizzas.json'
        if not pizzas_file.exists():
            self.skipTest(f"Datei nicht gefunden: {pizzas_file}")

        project = load_project(pizzas_file)

        self.assertIsInstance(project, Project)
        if project.total_volume is not None:
            self.assertGreater(project.total_volume, 0)
        if project.unit is not None:
            self.assertIsNotNone(project.unit)

    def test_load_loop_project(self):
        """Test: Projekt mit LoopTasks laden (erdaushub.json)"""
        erdaushub_file = _EXAMPLES_DIR / 'erdaushub.json'

        if not erdaushub_file.exists():
            self.skipTest(f"Datei nicht gefunden: {erdaushub_file}")

        project = load_project(erdaushub_file)

        self.assertIsInstance(project, Project)
        if project.total_volume is not None:
            self.assertGreater(project.total_volume, 0)
        if project.unit is not None:
            self.assertIsNotNone(project.unit)


class TestProjectMethods(unittest.TestCase):
    """Tests für Projekt-Methoden."""

    def test_get_time_unit(self):
        """Test: Zeiteinheit-Erkennung"""
        tankdesign_file = _EXAMPLES_DIR / 'tankdesign.json'

        if not tankdesign_file.exists():
            self.skipTest(f"Datei nicht gefunden: {tankdesign_file}")

        project = load_project(tankdesign_file)
        time_unit = project.get_time_unit()

        self.assertIn(time_unit, ['minutes', 'hours', 'days'])

    def test_get_all_task_ids(self):
        """Test: Alle Task-IDs abrufen"""
        tankdesign_file = _EXAMPLES_DIR / 'tankdesign.json'

        if not tankdesign_file.exists():
            self.skipTest(f"Datei nicht gefunden: {tankdesign_file}")

        project = load_project(tankdesign_file)
        task_ids = project.get_all_task_ids()

        self.assertEqual(len(task_ids), len(project.tasks))
        self.assertGreater(len(task_ids), 0)


class TestCPMCalculation(unittest.TestCase):
    """Tests für CPM-Berechnung."""

    def test_cpm_calculation_simple(self):
        """Test: CPM-Berechnung für einfaches Projekt"""
        tankdesign_file = _EXAMPLES_DIR / 'tankdesign.json'

        if not tankdesign_file.exists():
            self.skipTest(f"Datei nicht gefunden: {tankdesign_file}")

        project = load_project(tankdesign_file)
        result = project.calculate_cpm()

        self.assertIsNotNone(result)
        self.assertIsNotNone(result.project_duration)
        self.assertGreater(len(result.critical_path), 0)
        self.assertEqual(len(result.tasks), len(project.tasks))

    def test_cpm_critical_path(self):
        """Test: Kritischer Pfad wird korrekt identifiziert"""
        software_file = _EXAMPLES_DIR / 'software_simple.json'

        if not software_file.exists():
            self.skipTest(f"Datei nicht gefunden: {software_file}")

        project = load_project(software_file)
        result = project.calculate_cpm()

        self.assertGreater(len(result.critical_path), 0)

        for task_id in result.critical_path:
            task_result = result.tasks[task_id]
            self.assertAlmostEqual(task_result.puffer, 0.0, places=3)

    def test_cpm_with_start_date(self):
        """Test: CPM mit spezifischem Startdatum"""
        tankdesign_file = _EXAMPLES_DIR / 'tankdesign.json'

        if not tankdesign_file.exists():
            self.skipTest(f"Datei nicht gefunden: {tankdesign_file}")

        project = load_project(tankdesign_file)
        start_date = "2026-04-01"
        result = project.calculate_cpm(start_date=start_date)

        self.assertIsNotNone(result)
        self.assertIsNotNone(result.project_start)


class TestProjectExpansion(unittest.TestCase):
    """Tests für Projekt-Expansion (Zyklen, Loops)."""

    def test_expand_cycles(self):
        """Test: Zyklus-Expansion für Projekte mit InstanceTasks"""
        pizzas_file = _EXAMPLES_DIR / 'pizzas.json'

        if not pizzas_file.exists():
            self.skipTest(f"Datei nicht gefunden: {pizzas_file}")

        project = load_project(pizzas_file)

        if not any(isinstance(t, InstanceTask) for t in project.tasks):
            self.skipTest("Keine InstanceTasks im Projekt")

        original_task_count = len(project.tasks)
        expanded = project.expand_cycles()

        self.assertIsInstance(expanded, Project)
        self.assertGreaterEqual(len(expanded.tasks), original_task_count)

    def test_expand_loops(self):
        """Test: Loop-Expansion für Projekte mit LoopTasks"""
        erdaushub_file = _EXAMPLES_DIR / 'erdaushub.json'

        if not erdaushub_file.exists():
            self.skipTest(f"Datei nicht gefunden: {erdaushub_file}")

        project = load_project(erdaushub_file)

        if not any(isinstance(t, LoopTask) for t in project.tasks):
            self.skipTest("Keine LoopTasks im Projekt")

        original_task_count = len(project.tasks)
        expanded = project.expand_loops()

        self.assertIsInstance(expanded, Project)
        self.assertGreaterEqual(len(expanded.tasks), 0)


class TestProjectSaveLoad(unittest.TestCase):
    """Tests für Speichern und Laden von Projekten."""

    def test_save_and_load_project(self):
        """Test: Projekt speichern und wieder laden"""
        tankdesign_file = _EXAMPLES_DIR / 'tankdesign.json'

        if not tankdesign_file.exists():
            self.skipTest(f"Datei nicht gefunden: {tankdesign_file}")

        original = load_project(tankdesign_file)

        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            save_project(original, tmp_path)
            reloaded = load_project(tmp_path)

            self.assertEqual(original.project, reloaded.project)
            self.assertEqual(len(original.tasks), len(reloaded.tasks))

        finally:
            if tmp_path.exists():
                tmp_path.unlink()


class TestExcelExport(unittest.TestCase):
    """Tests für Excel-Export."""

    def test_excel_export_with_reports(self):
        """Test: Excel-Export mit Reports"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl nicht installiert")

        from lib.sproject import export_cpm_to_xlsx

        software_file = _EXAMPLES_DIR / 'software_simple.json'

        if not software_file.exists():
            self.skipTest(f"Datei nicht gefunden: {software_file}")

        project = load_project(software_file)
        result = project.calculate_cpm()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            cfg_dir = _CFG_DIR
            export_cpm_to_xlsx(result, tmp_path, project, cfg_dir)

            self.assertTrue(tmp_path.exists())
            self.assertGreater(tmp_path.stat().st_size, 0)

            wb = openpyxl.load_workbook(tmp_path)
            sheet_names = [ws.title for ws in wb.worksheets]

            self.assertGreaterEqual(len(sheet_names), 1)
            self.assertTrue(
                any(name in sheet_names for name in
                    ["Projektzusammenfassung", "Alle Tasks", "Kritischer Pfad", "CPM Analyse"]),
                f"Kein bekannter Zusammenfassungs-Tab gefunden in: {sheet_names}"
            )
            wb.close()

        finally:
            if tmp_path.exists():
                tmp_path.unlink()


class TestTxtExport(unittest.TestCase):
    """Tests für TXT-Export (inkl. netplan_table)."""

    def _make_project(self):
        from lib.models.tasks import SimpleTask
        return SimpleProject(
            project="Test TXT Export",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Task A", duration="3d"),
                SimpleTask(id=2, name="Task B", duration="2d", successors=[1]),
                SimpleTask(id=3, name="Task C", duration="4d", successors=[1]),
                SimpleTask(id=4, name="Task D", duration="1d", successors=[2, 3]),
            ]
        )

    def test_txt_export_creates_file(self):
        """Test: TXT-Export erzeugt Datei"""
        from lib.txt_export import export_cpm_to_txt

        project = self._make_project()
        result = project.calculate_cpm()

        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            success = export_cpm_to_txt(result, tmp_path, project_name=project.project)

            self.assertTrue(success)
            self.assertTrue(tmp_path.exists())
            self.assertGreater(tmp_path.stat().st_size, 0)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_txt_export_contains_all_sections(self):
        """Test: TXT-Export enthält alle konfigurierten Sektionen"""
        from lib.txt_export import export_cpm_to_txt

        project = self._make_project()
        result = project.calculate_cpm()

        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_cpm_to_txt(result, tmp_path, project_name=project.project)

            content = tmp_path.read_text(encoding='utf-8')

            self.assertIn("PROJEKTZUSAMMENFASSUNG", content)
            self.assertIn("KRITISCHER PFAD", content)
            self.assertIn("NETZPLAN TABELLE", content)
            self.assertIn("ALLE TASKS", content)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_txt_export_netplan_table_has_successors(self):
        """Test: netplan_table-Sektion enthält Nachfolger-Spalte"""
        from lib.txt_export import export_cpm_to_txt

        project = self._make_project()
        result = project.calculate_cpm()

        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_cpm_to_txt(result, tmp_path, project_name=project.project)

            content = tmp_path.read_text(encoding='utf-8')

            # Nachfolger-Spaltenheader muss in der NETZPLAN TABELLE Sektion erscheinen
            self.assertIn("Nachfolger", content)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_txt_export_from_example(self):
        """Test: TXT-Export mit echtem Beispielprojekt"""
        from lib.txt_export import export_cpm_to_txt

        tankdesign_file = _EXAMPLES_DIR / 'tankdesign.json'

        if not tankdesign_file.exists():
            self.skipTest(f"Datei nicht gefunden: {tankdesign_file}")

        project = load_project(tankdesign_file)
        result = project.calculate_cpm()
        cfg_dir = _CFG_DIR

        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            success = export_cpm_to_txt(
                result, tmp_path,
                project_name=project.project,
                project=project,
                cfg_dir=cfg_dir
            )

            self.assertTrue(success)
            content = tmp_path.read_text(encoding='utf-8')
            self.assertIn(project.project, content)

        finally:
            if tmp_path.exists():
                tmp_path.unlink()


class TestNetworkDiagram(unittest.TestCase):
    """Tests für Netzplan-Generierung."""

    def _make_result(self):
        from lib.models.tasks import SimpleTask
        project = SimpleProject(
            project="Network Test",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Start",    duration="2d"),
                SimpleTask(id=2, name="Branch A", duration="3d", successors=[1]),
                SimpleTask(id=3, name="Branch B", duration="5d", successors=[1]),
                SimpleTask(id=4, name="End",       duration="1d", successors=[2, 3]),
            ]
        )
        return project.calculate_cpm()

    def test_generate_network_csv_structure(self):
        """Test: to_network_csv erzeugt valides CSV"""
        result = self._make_result()
        csv_text = result.to_network_csv()

        lines = csv_text.splitlines()
        self.assertGreater(len(lines), 1)

        # Header-Zeile prüfen
        header = lines[0]
        self.assertIn("TaskID", header)
        self.assertIn("FAZ", header)
        self.assertIn("FEZ", header)
        self.assertIn("GP", header)
        self.assertIn("IsCritical", header)
        self.assertIn("Successors", header)

    def test_generate_network_csv_task_count(self):
        """Test: CSV enthält alle Tasks (ohne WE-Blocker)"""
        result = self._make_result()
        csv_text = result.to_network_csv()

        data_lines = [l for l in csv_text.splitlines()[1:] if l.strip()]
        non_we_tasks = [tid for tid in result.tasks
                        if not (isinstance(tid, str) and tid.startswith('WE-'))]
        self.assertEqual(len(data_lines), len(non_we_tasks))

    def test_generate_network_csv_critical_marker(self):
        """Test: Kritische Tasks werden als JA markiert"""
        result = self._make_result()
        csv_text = result.to_network_csv()

        # Mindestens eine JA-Markierung muss vorhanden sein
        self.assertIn("JA", csv_text)

    def test_generate_network_csv_no_we_blockers(self):
        """Test: WE-Blocker-Tasks werden nicht im CSV ausgegeben"""
        result = self._make_result()
        csv_text = result.to_network_csv()

        # WE-Blocker dürfen nicht im CSV erscheinen
        self.assertNotIn('"WE-', csv_text)


class TestPersonProject(unittest.TestCase):
    """Tests spezifisch für Projekte mit Personen."""

    def test_person_project_persons(self):
        """Test: Personen in Projekt mit persons-Feld"""
        software_file = _EXAMPLES_DIR / 'software_simple.json'

        if not software_file.exists():
            self.skipTest(f"Datei nicht gefunden: {software_file}")

        project = load_project(software_file)

        if not project.persons:
            self.skipTest("Projekt hat keine Personen")

        self.assertGreater(len(project.persons), 0)

        first_person = project.persons[0]
        self.assertIsNotNone(first_person.id)
        self.assertIsNotNone(first_person.name)
        # email ist Optional seit Modell-Vereinheitlichung
        if first_person.email is not None:
            self.assertIn('@', first_person.email)

        if hasattr(first_person, 'hourly_rate') and first_person.hourly_rate is not None:
            self.assertGreaterEqual(first_person.hourly_rate, 0)

    def test_person_project_resources(self):
        """Test: Ressourcen in Projekt mit resources-Feld"""
        software_file = _EXAMPLES_DIR / 'software_simple.json'

        if not software_file.exists():
            self.skipTest(f"Datei nicht gefunden: {software_file}")

        project = load_project(software_file)

        if not project.resources:
            self.skipTest("Projekt hat keine Ressourcen")

        self.assertGreater(len(project.resources), 0)

        first_resource = project.resources[0]
        self.assertIsNotNone(first_resource.id)
        self.assertIsNotNone(first_resource.name)


class TestTaskValidation(unittest.TestCase):
    """Tests für Task-Validierung."""

    def test_task_duration_validation(self):
        """Test: Dauer-Validierung"""
        from lib.models.tasks import SimpleTask

        valid_durations = ["10d", "5h", "30m", "2w"]
        for duration in valid_durations:
            task = SimpleTask(id=1, name="Test", duration=duration)
            self.assertEqual(task.duration, duration)

        task_numeric = SimpleTask(id=2, name="Test2", duration="5")
        self.assertIsNotNone(task_numeric.duration)

    def test_task_dependencies(self):
        """Test: Task-Abhängigkeiten (EA-Typ)"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(
            id=1,
            name="Test Task",
            duration="10d",
            successors=[2, 3]
        )

        self.assertEqual(len(task.successors), 2)
        self.assertIn(2, task.successors)
        self.assertIn(3, task.successors)

    def test_task_dependencies_aa(self):
        """Test: AA-Abhängigkeiten (Anfang-Anfang)"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(
            id=1,
            name="AA Task",
            duration="5d",
            successors_aa=[2]
        )

        self.assertEqual(len(task.successors_aa), 1)
        self.assertIn(2, task.successors_aa)

    def test_task_dependencies_ee(self):
        """Test: EE-Abhängigkeiten (Ende-Ende)"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(
            id=1,
            name="EE Task",
            duration="5d",
            successors_ee=[3]
        )

        self.assertEqual(len(task.successors_ee), 1)


class TestEdgeCasesEmptyProjects(unittest.TestCase):
    """Tests für Edge Cases: Leere und minimale Projekte."""

    def test_empty_task_list(self):
        """Test: Projekt mit leerer Task-Liste"""
        project = SimpleProject(
            project="Empty Project",
            project_start="2026-04-01",
            tasks=[]
        )

        self.assertEqual(len(project.tasks), 0)
        self.assertEqual(len(project.get_all_task_ids()), 0)

    def test_single_task_no_dependencies(self):
        """Test: Einzelner Task ohne Abhängigkeiten"""
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="Single Task Project",
            project_start="2026-04-01",
            tasks=[SimpleTask(id=1, name="Only Task", duration="5d")]
        )

        result = project.calculate_cpm()
        self.assertEqual(len(result.tasks), 1)
        self.assertIn(1, result.critical_path)
        if isinstance(result.project_duration, str):
            self.assertIn("5", result.project_duration)
        else:
            self.assertAlmostEqual(result.project_duration, 5.0, places=1)


class TestEdgeCasesInvalidDurations(unittest.TestCase):
    """Tests für Edge Cases: Ungültige Dauern."""

    def test_zero_duration_task(self):
        """Test: Task mit Dauer 0"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(id=1, name="Zero Duration", duration="0d")
        self.assertEqual(task.to_days(), 0.0)

    def test_very_large_duration(self):
        """Test: Task mit sehr großer Dauer"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(id=1, name="Long Task", duration="1000d")
        self.assertEqual(task.to_days(), 1000.0)

    def test_duration_with_different_units(self):
        """Test: Verschiedene Zeiteinheiten"""
        from lib.models.tasks import SimpleTask

        task_days    = SimpleTask(id=1, name="Days",    duration="10d")
        task_hours   = SimpleTask(id=2, name="Hours",   duration="8h")
        task_minutes = SimpleTask(id=3, name="Minutes", duration="480m")
        task_weeks   = SimpleTask(id=4, name="Weeks",   duration="2w")

        self.assertEqual(task_days.to_days(), 10.0)
        self.assertEqual(task_hours.to_days(), 1.0)
        self.assertEqual(task_minutes.to_days(), 1.0)
        self.assertEqual(task_weeks.to_days(), 10.0)

    def test_missing_duration_converts_to_zero(self):
        """Test: Fehlende Dauer wird als 0 behandelt"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(id=1, name="No Duration", duration=None)
        self.assertEqual(task.to_days(), 0.0)


class TestEdgeCasesCircularDependencies(unittest.TestCase):
    """Tests für zyklische Abhängigkeiten – CPM muss mit ValueError abbrechen."""

    def _make_project(self, tasks):
        from lib.models.tasks import SimpleTask
        return SimpleProject(
            project="Test",
            project_start="2026-04-01",
            tasks=[SimpleTask(**t) for t in tasks]
        )

    def test_self_dependency_raises(self):
        """Selbstreferenz (Task 1 -> 1) muss ValueError auslösen."""
        project = self._make_project([
            {"id": 1, "name": "Self Loop", "duration": "5d", "successors": [1]}
        ])
        with self.assertRaises(ValueError) as ctx:
            project.calculate_cpm()
        self.assertIn("Selbstreferenz", str(ctx.exception))
        self.assertIn("1", str(ctx.exception))

    def test_two_task_cycle_raises(self):
        """Direkter Zyklus (1 -> 2 -> 1) muss ValueError auslösen."""
        project = self._make_project([
            {"id": 1, "name": "Task A", "duration": "5d", "successors": [2]},
            {"id": 2, "name": "Task B", "duration": "5d", "successors": [1]},
        ])
        with self.assertRaises(ValueError) as ctx:
            project.calculate_cpm()
        msg = str(ctx.exception)
        self.assertIn("Zyklische Abhängigkeit", msg)
        self.assertIn("Task A", msg)
        self.assertIn("Task B", msg)

    def test_three_task_cycle_raises(self):
        """Indirekter Zyklus (1 -> 2 -> 3 -> 1) muss ValueError auslösen."""
        project = self._make_project([
            {"id": 1, "name": "Task A", "duration": "5d", "successors": [2]},
            {"id": 2, "name": "Task B", "duration": "5d", "successors": [3]},
            {"id": 3, "name": "Task C", "duration": "5d", "successors": [1]},
        ])
        with self.assertRaises(ValueError) as ctx:
            project.calculate_cpm()
        self.assertIn("Zyklische Abhängigkeit", str(ctx.exception))

    def test_cycle_in_subgraph_raises(self):
        """Zyklus in Teilgraph (Tasks 3->4->3), Rest azyklisch."""
        project = self._make_project([
            {"id": 1, "name": "Start",  "duration": "2d", "successors": [2]},
            {"id": 2, "name": "Middle", "duration": "3d", "successors": [3]},
            {"id": 3, "name": "CycleA", "duration": "4d", "successors": [4]},
            {"id": 4, "name": "CycleB", "duration": "4d", "successors": [3]},
        ])
        with self.assertRaises(ValueError) as ctx:
            project.calculate_cpm()
        self.assertIn("Zyklische Abhängigkeit", str(ctx.exception))

    def test_valid_project_no_error(self):
        """Fehlerfreier Graph (1->2->3) darf keinen ValueError auslösen."""
        project = self._make_project([
            {"id": 1, "name": "Task A", "duration": "5d", "successors": [2]},
            {"id": 2, "name": "Task B", "duration": "3d", "successors": [3]},
            {"id": 3, "name": "Task C", "duration": "2d", "successors": []},
        ])
        result = project.calculate_cpm()
        self.assertIsNotNone(result)
        self.assertEqual(result.project_duration, "10.0d")

    def test_software_simple_regression(self):
        """Regression: software_simple.json darf keine Zyklen haben."""
        software_file = _EXAMPLES_DIR / 'software_simple.json'
        if not software_file.exists():
            self.skipTest(f"Datei nicht gefunden: {software_file}")

        project = load_project(software_file)
        result = project.calculate_cpm()
        self.assertIsNotNone(result)
        for task_id, task in result.tasks.items():
            self.assertGreaterEqual(
                task.puffer, 0.0,
                f"Task {task_id} ('{task.name}') hat negativen Puffer: {task.puffer}"
            )


class TestEdgeCasesInvalidReferences(unittest.TestCase):
    """Tests für Edge Cases: Ungültige Referenzen."""

    def test_dependency_on_nonexistent_task(self):
        """Test: Abhängigkeit auf nicht existierenden Task"""
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="Missing Dependency",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Task 1", duration="5d", successors=[999])
            ]
        )

        try:
            result = project.calculate_cpm()
            self.assertIsNotNone(result)
        except Exception as e:
            self.assertIsNotNone(e)

    def test_duplicate_task_ids(self):
        """Test: Doppelte Task-IDs (tasks ist eine Liste → beide bleiben erhalten)"""
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="Duplicate IDs",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Task 1",           duration="5d"),
                SimpleTask(id=1, name="Task 1 Duplicate", duration="3d")
            ]
        )

        self.assertEqual(len(project.tasks), 2)


class TestEdgeCasesDateHandling(unittest.TestCase):
    """Tests für Edge Cases: Datumsbehandlung."""

    def test_invalid_date_format(self):
        """Test: Ungültiges Datumsformat wird als String gespeichert"""
        project = SimpleProject(
            project="Invalid Date",
            project_start="not-a-date",
            tasks=[]
        )
        self.assertEqual(project.project_start, "not-a-date")

    def test_date_without_time(self):
        """Test: Datum ohne Zeitangabe"""
        project = SimpleProject(
            project="Date Only",
            project_start="2026-04-01",
            tasks=[]
        )
        self.assertEqual(project.project_start, "2026-04-01")

    def test_date_with_time(self):
        """Test: Datum mit Zeitangabe"""
        project = SimpleProject(
            project="Date with Time",
            project_start="2026-04-01 09:00:00",
            tasks=[]
        )
        self.assertEqual(project.project_start, "2026-04-01 09:00:00")


class TestEdgeCasesCPMCalculation(unittest.TestCase):
    """Tests für Edge Cases: CPM-Berechnung."""

    def test_disconnected_task_groups(self):
        """Test: Unverbundene Task-Gruppen"""
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="Disconnected Groups",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Task A1", duration="5d"),
                SimpleTask(id=2, name="Task A2", duration="3d", successors=[1]),
                SimpleTask(id=3, name="Task B1", duration="4d"),
                SimpleTask(id=4, name="Task B2", duration="2d", successors=[3])
            ]
        )

        result = project.calculate_cpm()

        self.assertEqual(len(result.tasks), 4)
        self.assertGreater(len(result.critical_path), 0)

    def test_multiple_start_tasks(self):
        """Test: Mehrere Start-Tasks (keine Abhängigkeiten)"""
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="Multiple Starts",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Start 1", duration="5d", successors=[]),
                SimpleTask(id=2, name="Start 2", duration="3d", successors=[]),
                SimpleTask(id=3, name="End",     duration="2d", successors=[1, 2])
            ]
        )

        result = project.calculate_cpm()

        self.assertEqual(len(result.tasks), 3)

        if 3 in result.tasks:
            task3 = result.tasks[3]
            self.assertIsNotNone(task3.faz)
            self.assertGreaterEqual(task3.faz, 0)

    def test_long_critical_path(self):
        """Test: Langer kritischer Pfad (lineare Kette von 10 Tasks)"""
        from lib.models.tasks import SimpleTask

        tasks = []
        for i in range(1, 11):
            deps = [i - 1] if i > 1 else []
            tasks.append(SimpleTask(id=i, name=f"Task {i}", duration="1d", successors=deps))

        project = SimpleProject(
            project="Long Chain",
            project_start="2026-04-01",
            tasks=tasks
        )

        result = project.calculate_cpm()

        self.assertEqual(len(result.critical_path), 10)
        if isinstance(result.project_duration, str):
            self.assertIn("10", result.project_duration)
        else:
            self.assertAlmostEqual(result.project_duration, 10.0, places=1)


class TestEdgeCasesCycleExpansion(unittest.TestCase):
    """Tests für Edge Cases: Zyklus-Expansion."""

    def test_cycle_with_zero_order_volume(self):
        """Test: Zyklus mit Bestellvolumen 0"""
        from lib.models.resources import MachineResource
        from lib.models.tasks import InstanceTask

        project = Project(
            project="Zero Cycles",
            project_start="2026-04-01",
            order_volume=0,
            unit="Stück",
            resources=[MachineResource(id="R1", name="Resource 1", type="machine", unit="Stück/h")],
            tasks=[
                InstanceTask(id=1, name="Task", duration="1h", instances="order_volume")
            ]
        )

        expanded = project.expand_cycles()

        self.assertIsInstance(expanded, Project)

    def test_cycle_with_one_instance(self):
        """Test: Zyklus mit nur einer Instanz"""
        from lib.models.resources import MachineResource
        from lib.models.tasks import InstanceTask

        project = Project(
            project="Single Cycle",
            project_start="2026-04-01",
            order_volume=1,
            unit="Stück",
            resources=[MachineResource(id="R1", name="Resource 1", type="machine", unit="Stück/h")],
            tasks=[
                InstanceTask(id=1, name="Task", duration="1h", instances="order_volume")
            ]
        )

        expanded = project.expand_cycles()

        self.assertIsInstance(expanded, Project)
        self.assertGreater(len(expanded.tasks), 0)


class TestEdgeCasesLoopExpansion(unittest.TestCase):
    """Tests für Edge Cases: Loop-Expansion."""

    def test_loop_with_minimal_setup(self):
        """Test: Loop-Projekt mit minimaler Konfiguration (SimpleTask, keine LoopTasks)"""
        from lib.models.resources import MachineResource
        from lib.models.tasks import SimpleTask

        project = Project(
            project="Minimal Loop",
            project_start="2026-04-01",
            total_volume=100,
            unit="m3",
            resources=[MachineResource(id="R1", name="Resource 1", type="machine", capacity=10, unit="m3")],
            tasks=[
                SimpleTask(id=1, name="Prep", duration="1d")
            ]
        )

        self.assertEqual(project.total_volume, 100)
        self.assertEqual(len(project.tasks), 1)


class TestEdgeCasesResourceValidation(unittest.TestCase):
    """Tests für Edge Cases: Ressourcen-Validierung."""

    def test_negative_hourly_rate(self):
        """Test: Negativer Stundensatz wird akzeptiert (kein Validator vorhanden)"""
        from lib.models.resources import Person

        person = Person(
            id="test",
            name="Test Person",
            role="Developer",
            hourly_rate=-50.0
        )
        self.assertEqual(person.hourly_rate, -50.0)

    def test_zero_hourly_rate(self):
        """Test: Stundensatz 0 ist gültig"""
        from lib.models.resources import Person

        person = Person(
            id="intern",
            name="Intern",
            role="Intern",
            hourly_rate=0.0
        )
        self.assertEqual(person.hourly_rate, 0.0)

    def test_very_high_hourly_rate(self):
        """Test: Sehr hoher Stundensatz"""
        from lib.models.resources import Person

        person = Person(
            id="expert",
            name="Expert",
            role="Senior Consultant",
            hourly_rate=1000.0
        )
        self.assertEqual(person.hourly_rate, 1000.0)

    def test_email_is_optional(self):
        """Test: email-Feld ist Optional (kann None sein)"""
        from lib.models.resources import Person

        person_no_email = Person(id="p1", name="No Email", role="Dev", hourly_rate=80.0)
        self.assertIsNone(person_no_email.email)

        person_with_email = Person(
            id="p2", name="With Email", role="Dev",
            hourly_rate=80.0, email="dev@example.com"
        )
        self.assertEqual(person_with_email.email, "dev@example.com")


class TestEdgeCasesEmptyStrings(unittest.TestCase):
    """Tests für Edge Cases: Leere Strings und Whitespace."""

    def test_empty_project_name(self):
        """Test: Leerer Projektname"""
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="",
            project_start="2026-04-01",
            tasks=[SimpleTask(id=1, name="Task", duration="1d")]
        )
        self.assertEqual(project.project, "")

    def test_empty_task_name(self):
        """Test: Leerer Task-Name"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(id=1, name="", duration="1d")
        self.assertEqual(task.name, "")

    def test_whitespace_only_task_name(self):
        """Test: Task-Name mit nur Whitespace"""
        from lib.models.tasks import SimpleTask

        task = SimpleTask(id=1, name="   ", duration="1d")
        self.assertEqual(task.name, "   ")


class TestEdgeCasesCPMInfinityBug(unittest.TestCase):
    """Tests für Edge Cases: CPM sollte keine Infinity-Werte haben."""

    def test_no_infinity_in_cpm_results(self):
        """Test: CPM-Ergebnisse enthalten keine Infinity-Werte"""
        from lib.models.tasks import SimpleTask
        import math

        project = SimpleProject(
            project="No Infinity Test",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Task A", duration="1d"),
                SimpleTask(id=2, name="Task B", duration="2d", successors=[1]),
                SimpleTask(id=3, name="Task C", duration="1d", successors=[2])
            ]
        )

        result = project.calculate_cpm()

        for task_id, task_result in result.tasks.items():
            self.assertFalse(math.isinf(task_result.faz), f"Task {task_id} hat infinity in FAZ")
            self.assertFalse(math.isinf(task_result.fez), f"Task {task_id} hat infinity in FEZ")
            self.assertFalse(math.isinf(task_result.saz), f"Task {task_id} hat infinity in SAZ")
            self.assertFalse(math.isinf(task_result.sez), f"Task {task_id} hat infinity in SEZ")

    def test_single_task_no_infinity(self):
        """Test: Ein einzelner Task hat keine Infinity-Werte"""
        from lib.models.tasks import SimpleTask
        import math

        project = SimpleProject(
            project="Single Task",
            project_start="2026-04-01",
            tasks=[SimpleTask(id=1, name="Only Task", duration="5d")]
        )

        result = project.calculate_cpm()
        task_result = result.tasks[1]

        self.assertFalse(math.isinf(task_result.saz), "SAZ sollte nicht inf sein")
        self.assertFalse(math.isinf(task_result.sez), "SEZ sollte nicht inf sein")
        self.assertAlmostEqual(task_result.saz, task_result.faz, places=3)
        self.assertAlmostEqual(task_result.sez, task_result.fez, places=3)


class TestEdgeCasesComplexDependencies(unittest.TestCase):
    """Tests für Edge Cases: Komplexe Abhängigkeiten."""

    def test_task_with_many_dependencies(self):
        """Test: Task mit vielen Abhängigkeiten"""
        from lib.models.tasks import SimpleTask

        tasks = [SimpleTask(id=i, name=f"Task {i}", duration="1d") for i in range(1, 11)]
        tasks.append(SimpleTask(
            id=11,
            name="Final Task",
            duration="1d",
            successors=list(range(1, 11))
        ))

        project = SimpleProject(
            project="Many Dependencies",
            project_start="2026-04-01",
            tasks=tasks
        )

        result = project.calculate_cpm()

        self.assertEqual(len(result.tasks), 11)
        if 11 in result.tasks:
            final_task = result.tasks[11]
            self.assertIsNotNone(final_task.faz)
            self.assertGreaterEqual(final_task.faz, 0)

    def test_diamond_dependency_pattern(self):
        """Test: Diamant-Abhängigkeitsmuster (A -> B,C -> D)"""
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="Diamond Pattern",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Start",    duration="1d"),
                SimpleTask(id=2, name="Branch 1", duration="2d", successors=[1]),
                SimpleTask(id=3, name="Branch 2", duration="3d", successors=[1]),
                SimpleTask(id=4, name="Join",      duration="1d", successors=[2, 3])
            ]
        )

        result = project.calculate_cpm()

        # Branch 2 dauert länger → im kritischen Pfad
        self.assertIn(3, result.critical_path)
        self.assertIn(4, result.critical_path)


def _print_cpm_csv(result, title=""):
    """Hilfsfunktion: CPM-Ergebnis als CSV-Tabelle ausgeben."""
    if title:
        print(f"\n{'='*60}")
        print(f"  {title}")
        print(f"{'='*60}")
    print(result.to_network_csv())
    print(f"Projektdauer: {result.project_duration}  "
          f"Kritischer Pfad: {result.critical_path}")


class TestCPMDependencyTypes(unittest.TestCase):
    """
    Tests für alle vier Abhängigkeitstypen – je 3 Tasks pro Typ.

    Konvention in diesem Projekt:
      successors     = [Nachfolger-IDs]  (EA: Ende→Anfang, Standard)
      successors_aa  = [Nachfolger-IDs]  (AA: Anfang→Anfang, paralleler Start)
      successors_ee  = [Nachfolger-IDs]  (EE: Ende→Ende, gleichzeitiges Ende)
      successors_ae  = [Nachfolger-IDs]  (AE: Anfang→Ende, Nachfolger endet
                                                  wenn dieser startet)
    """

    @staticmethod
    def _build(name, tasks):
        return SimpleProject(project=name, project_start="2026-04-01", tasks=tasks)



    # EA	Sequenz: T2 startet wenn T1 endet	T1(0→5) → T2(5→8) → T3(8→10)
    def test_ea_simple_chain(self):
        """EA (Ende-Anfang): klassische Sequenz T1→T2→T3."""
        from lib.models.tasks import SimpleTask
        #
        #  T1(5d) ──EA──> T2(3d) ──EA──> T3(2d)
        #  FAZ=0  FEZ=5   FAZ=5  FEZ=8   FAZ=8  FEZ=10
        #  GP=0            GP=0            GP=0
        #  Kritischer Pfad: alle drei
        #
        project = self._build("EA-Test", [
            SimpleTask(id=1, name="Vorbereitung", duration="5d", successors=[2]),
            SimpleTask(id=2, name="Durchführung", duration="3d", successors=[3]),
            SimpleTask(id=3, name="Abschluss",    duration="2d", successors=[]),
        ])
        result = project.calculate_cpm()
        _print_cpm_csv(result, "EA (Ende-Anfang) – 3 Tasks")

        t1, t2, t3 = result.tasks[1], result.tasks[2], result.tasks[3]

        # Sequenz: jeder startet nach dem Ende des Vorgängers
        self.assertAlmostEqual(t1.faz, 0.0, places=3)
        self.assertAlmostEqual(t1.fez, 5.0, places=3)
        self.assertAlmostEqual(t2.faz, 5.0, places=3)
        self.assertAlmostEqual(t2.fez, 8.0, places=3)
        self.assertAlmostEqual(t3.faz, 8.0, places=3)
        self.assertAlmostEqual(t3.fez, 10.0, places=3)

        # Alle drei auf kritischem Pfad (GP=0)
        for tid in (1, 2, 3):
            self.assertAlmostEqual(result.tasks[tid].puffer, 0.0, places=3)
        self.assertEqual(sorted(result.critical_path), [1, 2, 3])

    # AA	Paralleler Start: T2 beginnt mit T1	T1(0→8) ∥ T2(0→3), beide enden vor T3(8→10). T2 hat GP=5
    def test_aa_parallel_start(self):
        """AA (Anfang-Anfang): T2 startet gleichzeitig mit T1."""
        from lib.models.tasks import SimpleTask
        #
        #  T1(8d) ──AA──> T2(3d)     beide starten bei FAZ=0
        #  T1     ──EA──> T3(2d)     T3 startet nach T1 (FAZ=8)
        #
        #  T1: FAZ=0, FEZ=8, GP=0  (kritisch: erzwingt T3-Start)
        #  T2: FAZ=0, FEZ=3, GP=5  (Puffer: 5d vor T3 abgeschlossen)
        #  T3: FAZ=8, FEZ=10, GP=0 (kritisch)
        #
        project = self._build("AA-Test", [
            SimpleTask(id=1, name="Aushub",       duration="8d",
                       successors=[3], successors_aa=[2]),
            SimpleTask(id=2, name="Abtransport",  duration="3d", successors=[3]),
            SimpleTask(id=3, name="Fundament",    duration="2d", successors=[]),
        ])
        result = project.calculate_cpm()
        _print_cpm_csv(result, "AA (Anfang-Anfang) – 3 Tasks")

        t1, t2, t3 = result.tasks[1], result.tasks[2], result.tasks[3]

        # T1 und T2 starten gleichzeitig (AA)
        self.assertAlmostEqual(t1.faz, t2.faz, places=3,
                               msg="AA: T1 und T2 starten zum gleichen Zeitpunkt")
        # T2 hat Puffer (kürzer als T1)
        self.assertGreater(t2.puffer, 0.0,
                           msg="AA: T2 (kürzer) hat positiven Puffer")
        # T3 folgt T1 (EA): T3.FAZ = T1.FEZ
        self.assertAlmostEqual(t3.faz, t1.fez, places=3,
                               msg="EA: T3 startet wenn T1 endet")
        # Kritischer Pfad: T1 und T3
        self.assertIn(1, result.critical_path)
        self.assertIn(3, result.critical_path)
        self.assertNotIn(2, result.critical_path)

    # EE	Gleichzeitiges Ende: T2.FEZ=T1.FEZ	T1(0→8) ∥ T2(5→8, verschoben), dann T3(8→10)
    def test_ee_simultaneous_end(self):
        """EE (Ende-Ende): T2 wird so verschoben, dass es gleichzeitig mit T1 endet."""
        from lib.models.tasks import SimpleTask
        #
        #  T1(8d) ──EE──> T2(3d)     T2.FEZ wird auf T1.FEZ=8 hochgezogen
        #  T1     ──EA──> T3(2d)     T3 startet nach T1 (FAZ=8)
        #
        #  T1: FAZ=0, FEZ=8, GP=0
        #  T2: FAZ=5, FEZ=8, GP=0  (von 0→3 auf 5→8 verschoben)
        #  T3: FAZ=8, FEZ=10, GP=0
        #
        project = self._build("EE-Test", [
            SimpleTask(id=1, name="Innenausbau",  duration="8d",
                       successors=[3], successors_ee=[2]),
            SimpleTask(id=2, name="Reinigung",    duration="3d", successors=[3]),
            SimpleTask(id=3, name="Abnahme",      duration="2d", successors=[]),
        ])
        result = project.calculate_cpm()
        _print_cpm_csv(result, "EE (Ende-Ende) – 3 Tasks")

        t1, t2, t3 = result.tasks[1], result.tasks[2], result.tasks[3]

        # T2.FEZ == T1.FEZ (EE-Constraint)
        self.assertAlmostEqual(t2.fez, t1.fez, places=3,
                               msg="EE: T2.FEZ muss gleich T1.FEZ sein")
        # T2 wurde nach hinten verschoben (FAZ > 0)
        self.assertGreater(t2.faz, 0.0,
                           msg="EE: T2.FAZ wurde von 0 nach hinten verschoben")
        # T3 folgt T1 (EA)
        self.assertAlmostEqual(t3.faz, t1.fez, places=3)
        # T2 ebenfalls auf kritischem Pfad (GP=0)
        self.assertAlmostEqual(t2.puffer, 0.0, places=3,
                               msg="EE: T2 ist kritisch (GP=0)")

    # AE	T2 endet wenn T1 startet	T3(0→5) → T1(5→10); T2(2→5) läuft parallel, endet bei T1.FAZ=5
    def test_ae_target_forced_to_end_before_predecessor_starts(self):
        """AE (Anfang-Ende): T2 muss enden bevor/wenn T1 startet."""
        from lib.models.tasks import SimpleTask
        #
        #  T1(5d) ──AE──> T2(3d)
        #  T0(5d) ──EA──> T1          T1 startet erst nach T0 (FAZ=5)
        #  T2 muss FEZ >= T1.FAZ=5  → T2.FAZ = 5-3 = 2, T2.FEZ = 5
        #
        #  T0: FAZ=0,     FEZ=5,         GP=0
        #  T1:            FAZ=5, FEZ=10, GP=0
        #  T2:     FAZ=2, FEZ=5,         GP=0  (parallel: läuft von 2→5 gleichzeitig mit Ende von T0)
        #
        project = self._build("AE-Test", [
            SimpleTask(id=1, name="Lieferung",   duration="5d",
                       successors=[], successors_ae=[2]),
            SimpleTask(id=2, name="Vorbereitung", duration="3d", successors=[]),
            SimpleTask(id=3, name="Vorgänger",    duration="5d", successors=[1]),
        ])
        result = project.calculate_cpm()
        _print_cpm_csv(result, "AE (Anfang-Ende) – 3 Tasks")

        t1, t2, t3 = result.tasks[1], result.tasks[2], result.tasks[3]

        # AE-Constraint: T2.FEZ >= T1.FAZ
        self.assertGreaterEqual(
            t2.fez, t1.faz - 0.001,
            msg="AE: T2.FEZ muss >= T1.FAZ sein (T2 endet wenn T1 startet)")

        # T2 wurde nach hinten verschoben (required_faz = T1.FAZ - dur(T2) > 0)
        expected_t2_faz = max(0.0, t1.faz - 3.0)
        self.assertAlmostEqual(t2.faz, expected_t2_faz, places=3,
                               msg="AE: T2.FAZ = max(0, T1.FAZ - dur(T2))")


class TestCPMResultExportToDict(unittest.TestCase):
    """Tests für CPMResult.export_to_dict() – Struktur und Inhalte."""

    def _make_result(self):
        from lib.models.tasks import SimpleTask
        # Einfaches 3-Task Projekt: T1→T2, T1→T3 (T1 ist Vorgänger beider)
        project = SimpleProject(
            project="Export Test",
            project_start="2026-06-01",
            tasks=[
                SimpleTask(id=1, name="Start",  duration="3d", successors=[2, 3]),
                SimpleTask(id=2, name="Zweig A", duration="2d", successors=[]),
                SimpleTask(id=3, name="Zweig B", duration="4d", successors=[]),
            ]
        )
        return project.calculate_cpm()

    def test_export_to_dict_top_level_keys(self):
        """export_to_dict enthält alle Pflicht-Schlüssel"""
        result = self._make_result()
        d = result.export_to_dict(include_dates=False)

        for key in ('project', 'project_duration', 'critical_path', 'tasks', 'settings'):
            self.assertIn(key, d, msg=f"Schlüssel '{key}' fehlt in export_to_dict")

    def test_export_to_dict_task_cpm_fields(self):
        """Jeder Task enthält alle CPM-Felder (D, FB, FE, SB, SE, GP, FP, is_critical)"""
        result = self._make_result()
        d = result.export_to_dict(include_dates=False)

        self.assertGreater(len(d['tasks']), 0)
        for task_dict in d['tasks']:
            self.assertIn('cpm', task_dict)
            for field in ('D', 'FB', 'FE', 'SB', 'SE', 'GP', 'FP', 'is_critical'):
                self.assertIn(field, task_dict['cpm'],
                              msg=f"CPM-Feld '{field}' fehlt in Task {task_dict.get('id')}")

    def test_export_to_dict_include_dates_false_no_date_block(self):
        """include_dates=False: kein 'dates'-Block in Task-CPM"""
        result = self._make_result()
        d = result.export_to_dict(include_dates=False)

        for task_dict in d['tasks']:
            self.assertNotIn('dates', task_dict.get('cpm', {}),
                             msg="'dates' darf nicht enthalten sein wenn include_dates=False")

    def test_export_to_dict_include_dates_true_has_date_block(self):
        """include_dates=True mit project_start: 'dates'-Block mit fb_date etc."""
        result = self._make_result()
        d = result.export_to_dict(include_dates=True)

        for task_dict in d['tasks']:
            self.assertIn('dates', task_dict['cpm'],
                          msg="'dates'-Block fehlt wenn include_dates=True")
            for key in ('fb_date', 'fe_date', 'sb_date', 'se_date'):
                self.assertIn(key, task_dict['cpm']['dates'],
                              msg=f"Datumsfeld '{key}' fehlt")

    def test_export_to_dict_task_count_matches(self):
        """Anzahl Tasks im Dict == Anzahl Tasks im CPMResult"""
        result = self._make_result()
        d = result.export_to_dict(include_dates=False)
        self.assertEqual(len(d['tasks']), len(result.tasks))

    def test_export_to_dict_critical_path_ids_valid(self):
        """critical_path enthält nur IDs die auch in tasks vorkommen"""
        result = self._make_result()
        d = result.export_to_dict(include_dates=False)

        self.assertGreater(len(d['critical_path']), 0,
                           msg="critical_path darf nicht leer sein")
        valid_ids = {t['id'] for t in d['tasks']}
        for cp_id in d['critical_path']:
            self.assertIn(cp_id, valid_ids,
                          msg=f"critical_path-ID {cp_id!r} nicht in tasks")


class TestNetworkCSVDetails(unittest.TestCase):
    """Tests für to_network_csv() – Successor-Spalte und Zahlenwerte."""

    def _make_result(self):
        from lib.models.tasks import SimpleTask
        #
        #  T1 ──EA──> T2 ──EA──> T4(end)
        #  T1 ──EA──> T3 ──EA──> T4(end)
        #
        #  successors zeigen auf Nachfolger:
        #  T1.deps=[2,3], T2.deps=[4], T3.deps=[4], T4.deps=[]
        #
        project = SimpleProject(
            project="CSV Detail Test",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Start",    duration="2d", successors=[2, 3]),
                SimpleTask(id=2, name="Branch A", duration="3d", successors=[4]),
                SimpleTask(id=3, name="Branch B", duration="5d", successors=[4]),
                SimpleTask(id=4, name="End",       duration="1d", successors=[]),
            ]
        )
        return project.calculate_cpm()

    def _parse_csv(self, csv_text):
        import csv, io
        return list(csv.DictReader(io.StringIO(csv_text), quotechar='"'))

    def test_multi_successor_semicolon_separated(self):
        """Task mit 2 Nachfolgern: Successors-Feld ist semikolongetrennt"""
        result = self._make_result()
        _print_cpm_csv(result, "CSV – Successor-Format")
        rows = self._parse_csv(result.to_network_csv())

        row1 = next(r for r in rows if r['TaskID'] == '1')
        succ = row1['Successors']
        self.assertIn(';', succ,
                      msg="Task 1 hat zwei Nachfolger, Trenner muss ';' sein")
        parts = set(succ.split(';'))
        self.assertEqual(parts, {'2', '3'})

    def test_end_task_successors_empty(self):
        """Task ohne Nachfolger (End): Successors-Feld ist leer"""
        result = self._make_result()
        rows = self._parse_csv(result.to_network_csv())

        row4 = next(r for r in rows if r['TaskID'] == '4')
        self.assertEqual(row4['Successors'].strip(), '',
                         msg="End-Task (deps=[]) muss leeres Successors-Feld haben")

    def test_single_successor_no_trailing_semicolon(self):
        """Task mit einem Nachfolger: kein abschließendes Semikolon"""
        result = self._make_result()
        rows = self._parse_csv(result.to_network_csv())

        for task_id in ('2', '3'):
            row = next(r for r in rows if r['TaskID'] == task_id)
            succ = row['Successors']
            self.assertFalse(succ.endswith(';'),
                             msg=f"Task {task_id}: kein trailing ';'")
            self.assertNotEqual(succ.strip(), '',
                                msg=f"Task {task_id}: muss Nachfolger haben")

    def test_no_inf_or_nan_in_numeric_columns(self):
        """Keine inf/nan in FAZ/FEZ/SAZ/SEZ/GP/FP"""
        import math
        result = self._make_result()
        rows = self._parse_csv(result.to_network_csv())

        for row in rows:
            for col in ('FAZ', 'FEZ', 'SAZ', 'SEZ', 'GP', 'FP'):
                val = float(row[col])
                self.assertFalse(math.isinf(val),
                                 msg=f"{col}=inf in Task {row['TaskID']}")
                self.assertFalse(math.isnan(val),
                                 msg=f"{col}=nan in Task {row['TaskID']}")


class TestExcelExportContent(unittest.TestCase):
    """Tests für Excel-Export: Sheet-Inhalte, nicht nur Namen."""

    def _make_project_and_result(self):
        from lib.models.tasks import SimpleTask
        project = SimpleProject(
            project="Excel Content Test",
            project_start="2026-04-01",
            tasks=[
                SimpleTask(id=1, name="Alpha", duration="3d", successors=[2, 3]),
                SimpleTask(id=2, name="Beta",  duration="2d", successors=[4]),
                SimpleTask(id=3, name="Gamma", duration="5d", successors=[4]),
                SimpleTask(id=4, name="Delta", duration="1d", successors=[]),
            ]
        )
        return project, project.calculate_cpm()

    def test_alle_tasks_sheet_has_enough_rows(self):
        """'Alle Tasks'-Sheet hat mindestens eine Zeile pro Task"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl nicht installiert")
        from lib.sproject import export_cpm_to_xlsx

        project, result = self._make_project_and_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            export_cpm_to_xlsx(result, tmp_path, project, _CFG_DIR)
            wb = openpyxl.load_workbook(tmp_path)

            task_sheet = next(
                (ws for ws in wb.worksheets
                 if any(kw in ws.title for kw in ("Alle", "Tasks", "Taskliste"))),
                wb.worksheets[0]
            )
            data_rows = [r for r in task_sheet.iter_rows(min_row=2, values_only=True)
                         if any(c is not None for c in r)]
            self.assertGreaterEqual(
                len(data_rows), len(project.tasks),
                msg=f"'Alle Tasks'-Sheet: erwartet >={len(project.tasks)} Zeilen, "
                    f"gefunden {len(data_rows)}")
            wb.close()
        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_kritischer_pfad_sheet_not_more_rows_than_critical_tasks(self):
        """'Kritischer Pfad'-Sheet: nicht mehr Datenzeilen als kritische Tasks"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl nicht installiert")
        from lib.sproject import export_cpm_to_xlsx

        project, result = self._make_project_and_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            export_cpm_to_xlsx(result, tmp_path, project, _CFG_DIR)
            wb = openpyxl.load_workbook(tmp_path)

            cp_sheet = next(
                (ws for ws in wb.worksheets
                 if any(kw in ws.title for kw in ("Kritisch", "Critical", "Pfad"))),
                None
            )
            if cp_sheet is None:
                self.skipTest("Kein Kritischer-Pfad-Sheet vorhanden")

            data_rows = [r for r in cp_sheet.iter_rows(min_row=2, values_only=True)
                         if any(c is not None for c in r)]
            # Toleranz +2 für mögliche Summen-/Trennzeilen
            self.assertLessEqual(
                len(data_rows), len(result.critical_path) + 2,
                msg="Kritischer-Pfad-Sheet enthält mehr Zeilen als kritische Tasks")
            wb.close()
        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_excel_export_single_task_no_crash(self):
        """Excel-Export mit Minimal-Projekt (1 Task): kein Absturz, Datei vorhanden"""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl nicht installiert")
        from lib.sproject import export_cpm_to_xlsx
        from lib.models.tasks import SimpleTask

        project = SimpleProject(
            project="Minimal", project_start="2026-04-01",
            tasks=[SimpleTask(id=1, name="Einziger Task", duration="1d")]
        )
        result = project.calculate_cpm()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            export_cpm_to_xlsx(result, tmp_path, project, _CFG_DIR)
            self.assertTrue(tmp_path.exists())
            self.assertGreater(tmp_path.stat().st_size, 0)
        finally:
            if tmp_path.exists():
                tmp_path.unlink()


def run_tests():
    """Führt alle Tests aus."""
    loader = unittest.TestLoader()
    suite = loader.loadTestsFromModule(sys.modules[__name__])
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    return 0 if result.wasSuccessful() else 1


if __name__ == '__main__':
    sys.exit(run_tests())
